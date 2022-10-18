import pandas as np
import os

class DataAnalysis:
    def __init__(self, location, output, modified_peptides, unmodified_peptides, internal_standard, template_file, debug):
        """ Object initialization.
            
            Parameters:
            ----------
            location: location of the file highest_intensities_per_raw csv file in wide format. 
                     tThis file is used to fill-put the template_file

            output: name of the excel file s(xlsx format) containing data analysis and plots 
            modified_peptides  : set of all modified peptides. Read from peak.ini file 
            unmodified_peptides: set of all unmodified_peptides. Read from peak.ini file
            internal_starndard : set of all internal standard. Read from peak.ini file
            template_file: location of the excel template file 'template.xlsx'. 
            debug:  optional parameter (boolean) for debbuging purposes.
                    set to False for default.
            
            exceptions:

        """
        # check that location file exist, otherwise throw an error
        assert os.path.exists(location), \
            f"{location} file does not exits.\nPLease Re-run peakEXtractor.py to generate it."
        
        # check that template file exist
        assert os.path.exists(template_file), \
            f"{template_file} template does not exits.\nPLease copy {template_file} from source code to current directory."
        
        self.location = location
        self.output   = output
        self.modified_peptides   = modified_peptides
        self.unmodified_peptides = unmodified_peptides
        self.internal_standard   = internal_standard
        self.template_file = template_file
        self.debug = debug

    def do_analysis(self):
        """
        Add description
        """
        import pandas as pd
        import os
        import openpyxl as xl

        csv_filename = self.location

        df = pd.read_csv(csv_filename)
        self.debug and print(df.head())

        # make a new column with empty values
        df['internal_s_total'] = pd.to_numeric(0)
        
        # add all internal standard columns in one column
        for i_standard in self.internal_standard:
            self.debug and print(f"internal_standard={i_standard}")
            header = f"{i_standard}-M/Z"
            # error checking, just in case M/Z is not found in the 
            # experimental data. 
            if header in df.columns:
                df['internal_s_total'] = df['internal_s_total'] + df[header]
        
        # make a new column with empty values
        df['modified_peptides'] = pd.to_numeric(0)
        
        # add all modified peptide-columns in one column
        for m in self.modified_peptides:
            header = f"{m}-M/Z"
            new_header = f"{header}-RATIO"
            # error checking, just in case M/Z is not found 
            # in the experimental data.
            if header in df.columns:
                df['modified_peptides'] = df['modified_peptides'] + df[header]
        
        # make a new column with empty values
        df['unmodified_peptides'] = pd.to_numeric(0)
        
        # add all modified peptides columns in one column
        for m in self.unmodified_peptides:
            header = f"{m}-M/Z"
            # error checking
            if header in df.columns:
                df['unmodified_peptides'] = df['unmodified_peptides'] + df[header]
        
        self.debug and print(df.head())
    
        sheet_name = 'Raw data'
        column_number = 1
        column_name = "SAMPLE"
        
        file_list = list()
        for file in df['SAMPLE']:
            self.debug and print(f"file={file} HERE")
            
            file_name = os.path.basename(file)
            self.debug and print(f"file_name ={file_name}")
            file_list.append(file_name)
        
        # opening an excel template file
        wb = xl.load_workbook(self.template_file) 
        self.debug and print(f"opening = {self.template_file}")
        
        self.update_excel_column_values(wb, sheet_name, column_number, column_name, df, file_list)   
        
        sheet_name = 'Raw data'
        column_number = 2
        column_name = "Angiotensin"
        self.update_excel_column_values(wb, sheet_name, column_number, column_name, df, df['internal_s_total'])   
        
        sheet_name = 'Raw data'
        column_number = 3
        column_name = "Modified peptide"
        self.update_excel_column_values(wb, sheet_name, column_number, column_name, df, df['modified_peptides'])   

        sheet_name = 'Raw data'
        column_number = 4
        column_name = "Unmodified peptide"
        self.update_excel_column_values(wb, sheet_name, column_number, column_name, df, df['unmodified_peptides'])   
        
        wb.save(self.output)

    @classmethod
    def update_excel_column_values(cls, wb, sheet_name, column_number, column_name, df, new_values):   
        import openpyxl as xl
        import itertools 
        debug = False

        #wb = xl.Workbook(path) 
        debug and print(wb.sheetnames)

        # check that sheet name exists
        debug and print(f"sheet_name={sheet_name}")
        sheet = wb[sheet_name]

        # count the number of rows and columns in sheet_name        
        row    = sheet.max_row
        column = sheet.max_column
        
        # count number of rows in the dataframe
        df_number_of_rows = len(df)
        debug and print(f'number of row={row}')
        debug and print(f'number of row={df_number_of_rows}')

        # delete column B
        sheet.delete_cols(column_number)
        # insert column B
        sheet.insert_cols(column_number)
        # and add column name
        sheet.cell(row=1, column=column_number).value=column_name
        
        # load the dataframe' data into a Python list
        data = []
        #for v in df['internal_s_total']:
        for v in new_values:
            data.append(v)
            debug and print(f"dataframe values {v}") 
        #debug and print(f"dimension = {len(data)}")
        
        # loop will print all values of column_number   
        #for i, v in zip( range(column_number, row + 1), df['internal_s_total'] ) :
        counter = 0
        start = 2
        stop = df_number_of_rows + 1
        for i in range(start, stop):
            cell_obj = sheet.cell(row = i, column = column_number)
            debug and print(f"column={column_number},row={i}, value={cell_obj.value} c={counter}")
            cell_obj.value = data[counter]
            counter += 1

if __name__ == '__main__':
    import os
    import time
    import sys
    import config_file as cfg
    from Raw import Raw 
    
    import pprint
    pp = pprint.PrettyPrinter(indent=4)
    DEBUG = False

    start = time.time()
    start_cpu_time = time.process_time()

    # read the configuration INI file and extract its values.
    ini_file = 'peak.ini'
    cfg_obj = cfg.config_file(location=ini_file, debug=DEBUG)
    values = cfg_obj.read_ini()

    DEBUG and pp.pprint(values)
    #pp.pprint(values)
    #sys.exit()
    idir = values['data_folder']
    odir = values['output']
    level = values['ms_level']
    ppm = values['ppm']
    debug = values['debug']
    masses = values['list_of_masses']
    i_standard = values['internal_standard']
    m_peptides = values['modified_peptides']
    u_peptides = values['unmodified_peptides']

    csv_filename='highest_intensities_per_raw_wide.csv'
    template_file = 'template.xlsx'
    output_plot_file = 'analysis_plot.xlsx'
    
    try:
        d = DataAnalysis(
            location = csv_filename, 
            modified_peptides  = m_peptides, 
            unmodified_peptides= u_peptides, 
            internal_standard  = i_standard,
            debug = debug,
            template_file = template_file,
            output = output_plot_file
        )
        d.do_analysis()
    except AssertionError as error:
        print(error)

    end = time.time()
    elapsed_time = end - start
    process_time = end - start_cpu_time
    print("=================================")
    print(f"Running time={elapsed_time} seconds")
    print(f'Wall time: {time.strftime("%H:%M:%S", time.gmtime(elapsed_time))}')
